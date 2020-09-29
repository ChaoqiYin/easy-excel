# !user/bin/env python3
# -*- coding: utf-8 -*-
# Author: ChaoqiYin
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side

from .base import Base, type_factory

BORDER_SIDE = Side(border_style='thin')

DEFAULT_STYLE = {
    'font': Font(name='Arial', size=10),
    'alignment': Alignment(horizontal='center', vertical='center', wrapText=True),  # 自动换行
    'border': Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)
}

DEFAULT_TITLE_STYLE = {
    'font': Font(name='Arial', size=14, bold=True),
    'alignment': Alignment(horizontal='center', vertical='center', wrapText=True),  # 自动换行
    'border': Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)
}


class Xlsx(Base):
    def __init__(self):
        super().__init__()
        self.wb = openpyxl.Workbook()
        # 删除本身的默认sheet
        for sheet in self.wb.worksheets:
            self.wb.remove(sheet)

    def save(self, file_path_or_stream):
        if isinstance(file_path_or_stream, str) and file_path_or_stream.find('.xlsx') < 0:
            file_path_or_stream += '.xlsx'
        self.wb.save(file_path_or_stream)
        self.wb.close()

    def get_sheet(self, sheet_name):
        return self._sheet_map.get(sheet_name)

    def add_sheet(self, sheet_name):
        self._sheet_map[sheet_name] = self.wb.create_sheet(title=sheet_name)

    def merge(self, sheet_name, start_row, end_row, start_col, end_col, style):
        sheet = self.get_sheet(sheet_name)
        sheet.merge_cells(start_row=start_row + 1, end_row=end_row + 1, start_column=start_col + 1, end_column=end_col + 1)

    def set_row_height(self, sheet_name, row_num, row_height):
        sheet = self.get_sheet(sheet_name)
        sheet.row_dimensions[row_num + 1].height = row_height  # 1为基准数

    def set_col_width(self, sheet_name, col_num, col_width):
        sheet = self.get_sheet(sheet_name)
        sheet.column_dimensions[get_column_letter(col_num + 1)].width = col_width  # 1为基准数

    @staticmethod
    def _get_style(style, is_title):
        # 有样式直接使用该样式，否则使用默认样式，区分是title样式或普通样式
        if style is not None:
            return style
        return DEFAULT_STYLE if is_title is False else DEFAULT_TITLE_STYLE

    def write(self, sheet_name, row_num, col_num, value, style, is_title=False):
        sheet = self.get_sheet(sheet_name)
        cell = sheet.cell(row_num + 1, col_num + 1, value)
        style = self._get_style(style, is_title)
        if isinstance(style, dict) is False:
            raise Exception('style must be dict, key is the cell style property name, value is the cell style')
        for style_name, style in style.items():
            setattr(cell, style_name, style)


type_factory['xlsx'] = Xlsx